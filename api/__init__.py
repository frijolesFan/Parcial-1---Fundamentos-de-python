# api/__init__.py (fragmento)

import unicodedata
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

#Crea la direccion con pathlib ya que genera problemas usando rutas convencionales

#Carpeta raíz del proyecto (dos niveles arriba desde api/)
BASE_DIR = Path(__file__).resolve().parent.parent

#Construye la ruta al archivo en data/
file_path = BASE_DIR / "data" / "resultado_laboratorio_suelo.xlsx"

#------------------------------------------------------------------------------

def _normalize_text(input_text: str) -> str:
    if input_text is None:
        return ""
    input_text = ''.join(c for c in unicodedata.normalize('NFKD', input_text) if not unicodedata.combining(c))
    return input_text.lower()

#------------------------------------------------------------------------------

def search_data(search_data: dict) -> list:
    """
    Busca datos en el archivo Excel basados en los criterios proporcionados.
    Sólo retorna las columnas: Departamento (B), Municipio (C), Cultivo (D) y Topografía (G).
    No limita la cantidad de resultados; recorre hasta el final del documento.
    """

    #Carga el libro y la hoja activa
    workbook = load_workbook(file_path)
    sheet = workbook.active

    #OBtiene encabezados de la primera fila
    headers = [cell.value for cell in sheet[1]]

    #Mapea índices de columnas relevantes si existen
    index_departamento = headers.index('Departamento') if 'Departamento' in headers else None
    index_municipio = headers.index('Municipio') if 'Municipio' in headers else None
    index_cultivo = headers.index('Cultivo') if 'Cultivo' in headers else None

    #Topografía puede venir con o sin acento, buscamos ambas
    topo_header = 'Topografía' if 'Topografía' in headers else ('Topografia' if 'Topografia' in headers else None)
    index_topografia = headers.index(topo_header) if topo_header else None

    #Criterios normalizados
    departamento_buscar = _normalize_text(search_data.get('Departamento', ''))
    municipio_buscar = _normalize_text(search_data.get('Municipio', ''))
    cultivo_buscar = _normalize_text(search_data.get('Cultivo', ''))

    # Límite solicitado
    raw_limit = search_data.get('Numero de datos')
    try:
        limit = int(raw_limit) if raw_limit is not None else None
        if limit is not None and limit < 0:
            limit = 0
    except (TypeError, ValueError):
        #Si el usuario ingresa algo no numérico devuelve todos los datos disponibles
        limit = None

    resultados = []

    #Itera en los datos y frena cuando no hay mas o cuando llega al tope
    for row in sheet.iter_rows(min_row=2, values_only=True):
        valor_departamento = _normalize_text(row[index_departamento]) if index_departamento is not None else ""
        valor_municipio = _normalize_text(row[index_municipio]) if index_municipio is not None else ""
        valor_cultivo = _normalize_text(row[index_cultivo]) if index_cultivo is not None else ""

        #Resultados filtrados
        if ((not departamento_buscar or departamento_buscar in valor_departamento) and
            (not municipio_buscar or municipio_buscar in valor_municipio) and
            (not cultivo_buscar or cultivo_buscar in valor_cultivo)):

            #Se almacenan las variables solicitadas capitalizadas
            fila_resultado = {}
            fila_resultado['Departamento'] = np.char.capitalize(row[index_departamento])
            fila_resultado['Municipio'] = np.char.capitalize(row[index_municipio])
            fila_resultado['Cultivo'] = np.char.capitalize(row[index_cultivo])
            fila_resultado[topo_header] = np.char.capitalize(row[index_topografia])

            #Almacena las variables edaficas
            fila_resultado['pH'] = row[11]      
            fila_resultado['Fosforo'] = row[13]
            fila_resultado['Potasio'] = row[19]

            resultados.append(fila_resultado)

            #Se detiene si hay menos valores que los solicitados
            if limit is not None and len(resultados) >= limit:
                break

    return resultados